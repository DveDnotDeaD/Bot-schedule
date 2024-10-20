# Тут библиотеки
from openpyxl import load_workbook
import openpyxl
import telebot
from telebot import types
import time
import random
import os
from dotenv import load_dotenv, find_dotenv

load_dotenv(find_dotenv())

# Необходимо ввести токен для бота
bot = telebot.TeleBot(os.getenv('TOKEN_TG'))

user_dict = {}


class User:
    def __init__(self, chat_id):
        self.chat_id = chat_id
        self.group_id = None
        self.last_message_time = 0


wb = load_workbook('./schedule.xlsx')

sheet = wb.active

groups = []
for row in wb.active['D3':'BF3']:
    for cell in row:
        if cell.value != None:
            groups.append(cell.value)

group_index = 0


def get_course_number_from_group_name(group_name):
    return int(group_name[0])


def get_courses_from_groups(groups):
    unique_courses = set()

    for group in groups:
        unique_courses.add(get_course_number_from_group_name(group))

    sorted_courses = sorted(list(unique_courses))

    return sorted_courses


days_of_week = []
for col in wb.active['B3':'B54']:
    for cell in col:
        if cell.value != None:
            days_of_week.append(cell.value)


def is_group_exist(maybe_group):
    flag = False

    for group in groups:
        if (maybe_group == group):
            flag = True
            break

    return flag


def is_day_exist(maybe_day):
    flag = False

    for day in days_of_week:
        if (maybe_day == day):
            flag = True
            break

    return flag


is_busy = False
delay_message = random.uniform(0, 1.2)
delay_user_message = 1.2


def get_group_schedule(group_id, day):
    day_index = days_of_week.index(day)
    lessons_number = []
    lesson = []

    for col in sheet.iter_cols(min_col=3, max_col=3, min_row=5+10*day_index, max_row=14+10*day_index, values_only=True):
        for num in col:
            lessons_number.append(num)

    for row in sheet.iter_rows(min_row=5+10*day_index, max_row=14+10*day_index, min_col=4+2*group_id, max_col=5 + 2*group_id, values_only=True):
        lesson.append(row)

    raw_schedule = dict(zip(lessons_number, lesson))

    filtered_schedule = {
        key: val for key,
        val in raw_schedule.items() if val != (None, None)
    }

    return filtered_schedule


courses = get_courses_from_groups(groups)


@bot.message_handler(commands=['start'])
def start(message):
    btn = types.KeyboardButton(text="Sorry for what?")
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    markup.add(btn)

    msg = bot.send_message(
        message.chat.id,
        text="Oh shit I'm sorry".format(message.from_user),
        reply_markup=markup
    )
    bot.register_next_step_handler(msg, course_step)


@bot.message_handler(func=lambda message: True, content_types=['text'])
def course_step(message):
    chat_id = message.from_user.id
    user = user_dict.get(chat_id, User(chat_id))
    user_dict[chat_id] = user
    current_time = time.time()
    if (current_time - user.last_message_time) < delay_user_message:
        bot.reply_to(
            message, 'Пожалуйста, подождите перед отправкой следующего сообщения.')
        return
    user.last_message_time = current_time
    btns = []

    for course in courses:
        btn = types.KeyboardButton(str(course) + " курс")
        btns.append(btn)

    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    markup.add(*btns)

    time.sleep(delay_message)
    msg = bot.send_message(
        chat_id,
        text="Выбери курс".format(message.from_user),
        reply_markup=markup
    )
    bot.register_next_step_handler(msg, group_step)


def group_step(message):
    chat_id = message.from_user.id
    user = user_dict.get(chat_id, User(chat_id))
    user_dict[chat_id] = user
    current_time = time.time()
    if (current_time - user.last_message_time) < delay_user_message:
        bot.reply_to(
            message, 'Пожалуйста, подождите перед отправкой следующего сообщения.')
        return
    user.last_message_time = current_time
    course = message.text

    if isinstance(course, str) and course[0] in list(map(str, courses)):
        global groups_in_course
        groups_in_course = [
            group for group in groups if group.startswith(course[0])]

        btns = []
        for group in groups_in_course:
            btn = types.KeyboardButton(group)
            btns.append(btn)

        back = types.KeyboardButton(text="Вернуться в главное меню")
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        markup.add(*btns, back)

        time.sleep(delay_message)

        msg = bot.send_message(
            chat_id,
            text="Выбери группу".format(message.from_user),
            reply_markup=markup
        )
        bot.register_next_step_handler(msg, day_step)
    else:
        msg = bot.send_message(
            chat_id,
            text="Выбери курс ещё раз"
        )
        bot.register_next_step_handler(msg, group_step)


def day_step(message):
    chat_id = message.chat.id
    user = user_dict.get(chat_id, User(chat_id))
    user_dict[chat_id] = user
    current_time = time.time()
    if (current_time - user.last_message_time) < delay_user_message:
        bot.reply_to(
            message, 'Пожалуйста, подождите перед отправкой следующего сообщения.')
        return
    user.last_message_time = current_time
    current_group = message.text

    if current_group in groups_in_course:
        global group_index
        group_index = groups.index(current_group)

        group_id = group_index
        user = user_dict[chat_id]
        user.group_id = group_id

        btns = []
        for day in days_of_week:
            btn = types.KeyboardButton(day)
            btns.append(btn)

        back = types.KeyboardButton("Вернуться в главное меню")

        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        markup.add(*btns, back)

        time.sleep(delay_message)

        msg = bot.send_message(
            chat_id,
            text="Выбери день недели".format(message.from_user),
            reply_markup=markup
        )
        bot.register_next_step_handler(msg, schedule_step)

    elif message.text == "Вернуться в главное меню":
        course_step(message)
    else:
        msg = bot.send_message(
            message.chat.id,
            text="Выбери группу мещё раз"
        )
        bot.register_next_step_handler(msg, day_step)


def schedule_step(message):
    chat_id = message.chat.id
    day = message.text
    user = user_dict[chat_id]
    if day in days_of_week:
        filtered_schedule = get_group_schedule(user.group_id, day)

        schedule_text = [f'{key} {value[0]} {value[1]} \n' for key,
                         value in filtered_schedule.items()]

        group_name = groups[user.group_id]

        message_text = group_name + "   " + day + "\n" + "".join(schedule_text)

        msg = bot.send_message(
            chat_id, text=message_text)

        bot.register_next_step_handler(msg, schedule_step)
    elif message.text == "Вернуться в главное меню":
        course_step(message)
    else:
        msg = bot.send_message(
            chat_id,
            text="Выбери день ещё раз"
        )
        bot.register_next_step_handler(msg, schedule_step)


bot.infinity_polling(timeout=10, long_polling_timeout=5)
